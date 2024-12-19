# Fecha
# No. Ck.
# A nombre de:
# RUC
# DV
# Monto
# ITBMS
# Total
# Concepto
# Proyecto
import re
from pathlib import Path

import openpyxl

from sage_templater.plugins.parsers.excel_parser import check_regular_expression, get_wb_and_sheets


def get_start_and_end_row_numbers(wb: openpyxl.Workbook, sheet_name: str) -> tuple[int, int]:
    """Get start and end row numbers from a sheet with checks and transfer format."""
    date_regexp = re.compile(r"\s*([Ff][Ee][cC][hH][aA])\s*")
    check_number_regexp = re.compile(r"\s*[Nn][Oo]\.?\s[Cc][Kk]\.?\s*")
    sheet = wb[sheet_name]
    start_row = -1
    if sheet.max_row is None:
        return start_row, -1
    end_row = sheet.max_row
    i = 0
    for row in sheet.iter_rows():
        i += 1
        date_cell_value = row[0].value
        check_number_value = row[1].value
        if date_cell_value is None or not isinstance(date_cell_value, str):
            continue
        date_match = check_regular_expression(date_regexp, date_cell_value)
        check_number_match = check_regular_expression(check_number_regexp, check_number_value)
        if date_match and check_number_match:
            start_row = i
            break
    return start_row, end_row


def is_check_template(excel_file: Path) -> bool:
    if excel_file.suffix != ".xlsx":
        return False
    wb, sheets = get_wb_and_sheets(excel_file)
    start, end = get_start_and_end_row_numbers(wb, sheets[0])
    return start != -1
